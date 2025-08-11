import customtkinter as ctk
import sqlite3
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import messagebox
from tkcalendar import Calendar
import csv
from tkinter import filedialog
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
import os


def construir_dia_administrativo(frame_padre):
    # --- FUNCIÓN PARA SELECCIÓN DE FECHA ---
    def seleccionar_fecha(entry_target):
        top = tk.Toplevel()
        cal = Calendar(top, date_pattern='dd/mm/yyyy', locale='es_CL')
        cal.pack(padx=10, pady=10)
        def poner_fecha():
            entry_target.delete(0, tk.END)
            entry_target.insert(0, cal.get_date())
            top.destroy()
            # Recalcular si corresponde
            if entry_target is entry_fecha_desde and auto_calc_var.get():
                actualizar_fecha_hasta_auto()
        tk.Button(top, text="Seleccionar", command=poner_fecha).pack(pady=5)

    def exportar_informe():
        rut = entry_rut.get().strip()
        if not rut:
            messagebox.showwarning("Falta RUT", "Primero selecciona/busca un trabajador (RUT).")
            return

        # ¿Exportar SOLO lo seleccionado?
        exportar_solo_seleccion = False
        if selected_ids:
            exportar_solo_seleccion = messagebox.askyesno(
                "Exportar",
                "¿Exportar SOLO los registros seleccionados?\n\nPulsa 'No' para exportar todos los del trabajador."
            )

        # Rango opcional si ambas fechas son válidas
        fi = entry_fecha_desde.get().strip()
        ff = entry_fecha_hasta.get().strip()
        fecha_ini = fecha_fin = None
        rango_legible = ""
        if fi and ff:
            try:
                fecha_ini = datetime.strptime(fi, "%d/%m/%Y").strftime("%Y-%m-%d")
                fecha_fin = datetime.strptime(ff, "%d/%m/%Y").strftime("%Y-%m-%d")
                rango_legible = f"{fi} a {ff}"
            except ValueError:
                fecha_ini = fecha_fin = None
                rango_legible = ""

        # Consultar filas + nombre funcionario
        try:
            conn = sqlite3.connect("reloj_control.db")
            cur = conn.cursor()

            if exportar_solo_seleccion:
                placeholders = ",".join("?" for _ in selected_ids)
                cur.execute(f"""
                    SELECT id, fecha, motivo
                    FROM dias_libres
                    WHERE id IN ({placeholders})
                    ORDER BY fecha
                """, tuple(selected_ids))
            else:
                sql = "SELECT id, fecha, motivo FROM dias_libres WHERE rut = ?"
                params = [rut]
                if fecha_ini and fecha_fin:
                    sql += " AND fecha BETWEEN ? AND ?"
                    params.extend([fecha_ini, fecha_fin])
                sql += " ORDER BY fecha"
                cur.execute(sql, params)

            filas = cur.fetchall()

            # Nombre del funcionario
            cur.execute("SELECT nombre, apellido FROM trabajadores WHERE rut = ?", (rut,))
            row = cur.fetchone()
            nombre_funcionario = (f"{row[0] or ''} {row[1] or ''}".strip()) if row else ""
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo consultar los datos:\n{e}")
            return

        if not filas:
            messagebox.showinfo("Sin datos", "No hay registros para exportar con los filtros actuales.")
            return

        # Nombre sugerido único (timestamp)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = f"informe_{rut}_{ts}"
        if exportar_solo_seleccion:
            base += f"_sel{len(filas)}"

        # Elegir destino
        ruta = filedialog.asksaveasfilename(
            title="Guardar informe",
            defaultextension=".xlsx",
            filetypes=[("Excel (*.xlsx)", "*.xlsx"), ("CSV (*.csv)", "*.csv")],
            initialfile=f"{base}.xlsx"
        )
        if not ruta:
            return

        # Si eligen CSV o no hay openpyxl → CSV con ;
        if ruta.lower().endswith(".csv") or Workbook is None:
            if Workbook is None and ruta.lower().endswith(".xlsx"):
                messagebox.showinfo(
                    "Sin openpyxl",
                    "No encontré 'openpyxl'. Exportaré en CSV con separador ';' para Excel."
                )
                ruta = ruta.rsplit(".", 1)[0] + ".csv"
            _exportar_csv(ruta, filas, rut)
            return

        # Excel bonito. Intentamos con la versión nueva (con resumen mejorado);
        # si tu _exportar_xlsx aún tiene la firma vieja, hacemos fallback.
        try:
            _exportar_xlsx(
                ruta, filas, rut,
                nombre_funcionario=nombre_funcionario,
                rango_legible=rango_legible,
                solo_seleccion=exportar_solo_seleccion
            )
        except TypeError:
            # Fallback si no actualizaste la firma de _exportar_xlsx
            _exportar_xlsx(ruta, filas, rut)



    def _ruta_unica(path: str) -> str:
        import os
        base, ext = os.path.splitext(path)
        i = 1
        out = path
        while os.path.exists(out):
            out = f"{base} ({i}){ext}"
            i += 1
        return out


    def _categoria_motivo(motivo: str) -> str:
        s = (motivo or "").lower()
        if "día administrativo" in s:
            return "Día Administrativo"
        if "matrimonio" in s or "unión civil" in s:
            return "Matrimonio/AUC"
        if "defunción" in s:
            if "hijo en gestación" in s:
                return "Defunción: Hijo en gestación"
            if "hijo" in s:
                return "Defunción: Hijo"
            if "cónyuge" in s or "conviviente civil" in s:
                return "Defunción: Cónyuge/Conviviente"
            if "padre" in s or "madre" in s or "hermano" in s:
                return "Defunción: Padre/Madre/Hermano(a)"
            return "Defunción: Otro"
        if "nacimiento paternal" in s:
            return "Nacimiento Paternal"
        if "alimentación" in s:
            return "Permiso de Alimentación"
        if "sin goce" in s:
            return "Permiso sin Goce de Sueldo"
        if "cometido de servicio" in s:
            return "Cometido de Servicio"
        if "licencia" in s:
            return "Licencia Médica"
        return "Otros permisos"


    def _exportar_csv(ruta, filas, rut):
        import csv
        try:
            with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f, delimiter=';')
                w.writerow(["RUT", "ID", "Fecha", "Motivo"])
                for id_, fecha_iso, motivo in filas:
                    fecha_legible = datetime.strptime(fecha_iso, "%Y-%m-%d").strftime("%d/%m/%Y")
                    w.writerow([rut, id_, fecha_legible, motivo or ""])
            messagebox.showinfo("Exportación exitosa", f"Informe CSV exportado en:\n{ruta}")
        except Exception as e:
            messagebox.showerror("Error al exportar CSV", str(e))


    def _exportar_xlsx(ruta, filas, rut, nombre_funcionario=None, rango_legible="", solo_seleccion=False):
        """
        Exporta a Excel con:
        - Hoja 'Detalle' (datos con formato, autofiltro, encabezado fijo)
        - Panel 'Resumen' al costado (cols F:G) con métricas clave
        - Hoja 'Resumen' con dos tablas: por categoría y por motivo
        """
        try:
            from collections import Counter

            wb = Workbook()

            # -------- Hoja Detalle --------
            ws = wb.active
            ws.title = "Detalle"
            headers = ["RUT", "ID", "Fecha", "Motivo"]
            ws.append(headers)

            # Contadores
            motivo_counts = Counter()
            categoria_counts = Counter()

            for id_, fecha_iso, motivo in filas:
                fecha_dt = datetime.strptime(fecha_iso, "%Y-%m-%d")
                ws.append([rut, id_, fecha_dt, motivo or ""])
                motivo_counts[motivo or ""] += 1
                categoria_counts[_categoria_motivo(motivo)] += 1

            # Estilos encabezado
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_font = Font(color="FFFFFF", bold=True)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Formato de fecha
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=3).number_format = "DD/MM/YYYY"

            # Autofiltro y fijar encabezado
            ws.auto_filter.ref = f"A1:D{ws.max_row}"
            ws.freeze_panes = "A2"

            # Ajuste de anchos aproximado
            from openpyxl.utils import get_column_letter
            col_max = [len(h) for h in headers]
            for row in range(2, ws.max_row + 1):
                v1 = str(ws.cell(row=row, column=1).value or "")
                v2 = str(ws.cell(row=row, column=2).value or "")
                v3v = ws.cell(row=row, column=3).value
                v3 = v3v.strftime("%d/%m/%Y") if isinstance(v3v, datetime) else (str(v3v) if v3v else "")
                v4 = str(ws.cell(row=row, column=4).value or "")
                for i, v in enumerate([v1, v2, v3, v4]):
                    col_max[i] = max(col_max[i], len(v))
            for i, width in enumerate(col_max, start=1):
                ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 60)

            # -------- Panel de Resumen al costado (col F:G) --------
            start_col = 6  # F
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col+1)
            titulo = ws.cell(row=1, column=start_col, value="Resumen")
            titulo.fill = header_fill
            titulo.font = header_font
            titulo.alignment = Alignment(horizontal="center", vertical="center")

            def _put_pair(r, label, value):
                ws.cell(row=r, column=start_col, value=label).alignment = Alignment(horizontal="left")
                ws.cell(row=r, column=start_col+1, value=value).alignment = Alignment(horizontal="right")

            fila = 2
            if nombre_funcionario:
                _put_pair(fila, "Funcionario", nombre_funcionario); fila += 1

            _put_pair(fila, "RUT", rut); fila += 1
            _put_pair(fila, "Rango aplicado", (rango_legible or "Todos")); fila += 1
            _put_pair(fila, "Solo seleccionados", "Sí" if solo_seleccion else "No"); fila += 1

            limite_admin = 6
            usados_admin = categoria_counts.get("Día Administrativo", 0)
            restantes_admin = max(0, limite_admin - usados_admin)
            total = sum(motivo_counts.values())
            licencias = categoria_counts.get("Licencia Médica", 0)
            extras = total - usados_admin  # (licencias se muestra aparte)

            resumen_pairs = [
                ("Total registros", total),
                ("Días administrativos (usados)", usados_admin),
                ("Límite anual", limite_admin),
                ("Restantes", restantes_admin),
                ("Permisos extras", extras),
                ("Licencias médicas", licencias),
            ]
            for etiqueta, valor in resumen_pairs:
                _put_pair(fila, etiqueta, valor)
                fila += 1

            # Subtítulo categorías
            fila += 1
            ws.merge_cells(start_row=fila, start_column=start_col, end_row=fila, end_column=start_col+1)
            sub = ws.cell(row=fila, column=start_col, value="Por categoría")
            sub.fill = PatternFill("solid", fgColor="D9E1F2")
            sub.font = Font(bold=True)
            sub.alignment = Alignment(horizontal="center")
            fila += 1

            for cat, cant in sorted(categoria_counts.items(), key=lambda x: (-x[1], x[0])):
                _put_pair(fila, cat, cant)
                fila += 1

            # Anchos del panel
            ws.column_dimensions[get_column_letter(start_col)].width = 32   # F
            ws.column_dimensions[get_column_letter(start_col+1)].width = 14 # G

            # -------- Hoja Resumen (2 tablas) --------
            resumen = wb.create_sheet("Resumen")

            # Tabla 1: por categoría
            resumen.append(["Categoría", "Días"])
            resumen.cell(row=1, column=1).fill = header_fill
            resumen.cell(row=1, column=2).fill = header_fill
            resumen.cell(row=1, column=1).font = header_font
            resumen.cell(row=1, column=2).font = header_font
            for cat, cant in sorted(categoria_counts.items(), key=lambda x: (-x[1], x[0])):
                resumen.append([cat, cant])
            resumen.auto_filter.ref = f"A1:B{resumen.max_row}"
            resumen.freeze_panes = "A2"
            resumen.column_dimensions["A"].width = 45
            resumen.column_dimensions["B"].width = 12

            # Tabla 2: por motivo (debajo, dejando una fila vacía)
            inicio2 = resumen.max_row + 2
            resumen.cell(row=inicio2, column=1, value="Motivo").fill = header_fill
            resumen.cell(row=inicio2, column=2, value="Días").fill = header_fill
            resumen.cell(row=inicio2, column=1).font = header_font
            resumen.cell(row=inicio2, column=2).font = header_font

            r = inicio2 + 1
            for mot, cant in sorted(motivo_counts.items(), key=lambda x: (-x[1], x[0])):
                resumen.cell(row=r, column=1, value=mot)
                resumen.cell(row=r, column=2, value=cant)
                r += 1

            wb.save(ruta)
            messagebox.showinfo("Exportación exitosa", f"Informe Excel exportado en:\n{ruta}")
        except Exception as e:
            messagebox.showerror("Error al exportar Excel", str(e))





    # --- CONFIGURACIÓN DE PERMISOS Y DÍAS ---
    opciones_permiso = [
        "Elija tipo de Solicitud o Permiso",
        "Día Administrativo",
        "Permiso por Matrimonio/Acuerdo Unión Civil (5 días hábiles)",
        "Permiso por Defunción de Hijo (10 días corridos)",
        "Permiso por Defunción de Cónyuge/Conviviente Civil (7 días corridos)",
        "Permiso por Defunción de Hijo en Gestación (7 días hábiles)",
        "Permiso por Defunción de Padre/Madre/Hermano(a) (4 días hábiles)",
        "Permiso de Nacimiento Paternal (5 días corridos)",
        "Permiso de Alimentación (1 hora diaria)",
        "Permiso sin Goce de Sueldo (máx 6 meses)",
        "Cometido de Servicio",
        "Otro (especificar)"
    ]
    dias_por_permiso = {
        "Permiso por Matrimonio/Acuerdo Unión Civil (5 días hábiles)": 5,
        "Permiso por Defunción de Hijo (10 días corridos)": 10,
        "Permiso por Defunción de Cónyuge/Conviviente Civil (7 días corridos)": 7,
        "Permiso por Defunción de Hijo en Gestación (7 días hábiles)": 7,
        "Permiso por Defunción de Padre/Madre/Hermano(a) (4 días hábiles)": 4,
        "Permiso de Nacimiento Paternal (5 días corridos)": 5,
        "Permiso sin Goce de Sueldo (máx 6 meses)": 180
    }

    # --- FRAME PRINCIPAL ---
    frame = ctk.CTkFrame(frame_padre)
    frame.pack(fill="both", expand=True, padx=20, pady=20)

    

    ctk.CTkLabel(frame, text="Asignar, Ver o Editar Días Administrativos, Permisos y Licencias", font=("Arial", 18)).pack(pady=10)
    

    # --- CARGA NOMBRES/RUTS Y AUTOCOMPLETADO ---
    def cargar_nombres_ruts():
        conn = sqlite3.connect("reloj_control.db")
        cursor = conn.cursor()
        cursor.execute("SELECT rut, nombre, apellido FROM trabajadores")
        nombres, dict_nombre_rut = [], {}
        for rut, nombre, apellido in cursor.fetchall():
            nombre_completo = f"{nombre} {apellido}".strip()
            nombres.append(nombre_completo)
            dict_nombre_rut[nombre_completo] = rut
        conn.close()
        return nombres, dict_nombre_rut

    lista_nombres, dict_nombre_rut = cargar_nombres_ruts()
    primeros_10 = lista_nombres[:10]

    # --- COMBOBOX Y BOTÓN DE BÚSQUEDA ---
    fila_nombre = ctk.CTkFrame(frame, fg_color="transparent")
    fila_nombre.pack(pady=10)
    combo_nombre = ctk.CTkComboBox(fila_nombre, values=primeros_10, width=250)
    combo_nombre.set("Buscar por Nombre")
    combo_nombre.pack(side="left", padx=(0, 5))
    btn_buscar = ctk.CTkButton(fila_nombre, text="🔍 Buscar", command=lambda: buscar_solicitudes())
    btn_buscar.pack(side="left", padx=(5, 0))

    # --- ENTRY DE RUT ---
    entry_rut = ctk.CTkEntry(frame, placeholder_text="RUT: (Ej: 12345678-9)", width=150)
    entry_rut.pack(pady=5)
    entry_rut.bind("<Return>", lambda event: mostrar_vista_previa())

    # --- LABELS PARA RESUMEN ---
    label_resumen_admin = ctk.CTkLabel(frame, text="", text_color="green", font=("Arial", 13))
    label_resumen_admin.pack(pady=(0, 5))
    label_resumen_extras = ctk.CTkLabel(frame, text="", text_color="red", font=("Arial", 13))
    label_resumen_extras.pack(pady=(0, 10))

    # --- SUBTITULO SECCIÓN DE NUEVA SOLICITUD ---
    ctk.CTkLabel(
        frame, 
        text="IMPORTANTE",
        font=("Arial", 13, "bold"),
        text_color="#d32f2f"
    ).pack(pady=(12, 2))
    ctk.CTkLabel(
        frame, 
        text="Nueva Solicitud de Permiso o Licencia   ▼",
        font=("Arial", 15, "bold"),
        text_color="#607D8B",
        anchor="center"
    ).pack(pady=(0, 8))

    # --- PERMISO (COMBOBOX) ---
    combo_tipo_permiso = ctk.CTkComboBox(frame, values=opciones_permiso, width=400)
    auto_calc_var = tk.BooleanVar(value=True)
    chk_auto = ctk.CTkCheckBox(frame, text="Calcular automáticamente la fecha hasta", variable=auto_calc_var)
    chk_auto.pack(pady=(2, 8))
    combo_tipo_permiso.set("Elija tipo de Solicitud o Permiso")
    combo_tipo_permiso.pack(pady=5)

    # --- FECHAS DESDE/HASTA ---
    cont_desde = ctk.CTkFrame(frame, fg_color="transparent")
    cont_desde.pack(pady=2)
    entry_fecha_desde = ctk.CTkEntry(cont_desde, placeholder_text="Fecha Desde (dd/mm/aaaa)", width=200)
    entry_fecha_desde.pack(side="left")
    ctk.CTkButton(cont_desde, text="📅", width=40, command=lambda: seleccionar_fecha(entry_fecha_desde)).pack(side="left", padx=5)

    cont_hasta = ctk.CTkFrame(frame, fg_color="transparent")
    cont_hasta.pack(pady=2)
    entry_fecha_hasta = ctk.CTkEntry(cont_hasta, placeholder_text="Fecha Hasta (dd/mm/aaaa)", width=200)
    entry_fecha_hasta.pack(side="left")
    ctk.CTkButton(cont_hasta, text="📅", width=40, command=lambda: seleccionar_fecha(entry_fecha_hasta)).pack(side="left", padx=5)
    if auto_calc_var.get():
        entry_fecha_hasta.configure(state="disabled")

    # --- MOTIVO EXTRA Y LABEL INFO ---
    entry_motivo_otro = ctk.CTkEntry(frame, placeholder_text="Especificar motivo...", width=400)
    entry_motivo_otro.pack(pady=(0, 10))
    entry_motivo_otro.pack_forget()
    label_info_cometido = ctk.CTkLabel(frame, text="", text_color="orange", font=("Arial", 12))
    label_info_cometido.pack()

    # ========== FUNCIONES DE AUTOCOMPLETADO Y BÚSQUEDA ==========
    def limpiar_placeholder(event):
        if combo_nombre.get() == "Buscar por Nombre":
            combo_nombre.set("")
    def restaurar_placeholder(event):
        if combo_nombre.get() == "":
            combo_nombre.set("Buscar por Nombre")
    combo_nombre.bind("<FocusIn>", limpiar_placeholder)
    combo_nombre.bind("<FocusOut>", restaurar_placeholder)

    def mostrar_sugerencias(event):
        texto = combo_nombre.get().lower()
        if not texto or texto == "buscar por nombre":
            combo_nombre.configure(values=primeros_10)
    combo_nombre.bind("<FocusIn>", mostrar_sugerencias)

    def autocompletar_nombres(event):
        texto = combo_nombre.get().lower()
        if not texto or texto == "buscar por nombre":
            combo_nombre.configure(values=primeros_10)
        else:
            filtrados = [n for n in lista_nombres if texto in n.lower()]
            combo_nombre.configure(values=filtrados[:10] if filtrados else ["No encontrado"])
    combo_nombre.bind("<KeyRelease>", autocompletar_nombres)

    def buscar_por_nombre(event=None):
        nombre = combo_nombre.get()
        rut = dict_nombre_rut.get(nombre, "")
        if rut:
            entry_rut.delete(0, tk.END)
            entry_rut.insert(0, rut)
            entry_fecha_desde.delete(0, tk.END)
            entry_fecha_hasta.delete(0, tk.END)
            mostrar_vista_previa()
        else:
            messagebox.showwarning("Nombre no válido", "Selecciona un nombre válido.")
    combo_nombre.bind("<Return>", buscar_por_nombre)
    combo_nombre.bind("<<ComboboxSelected>>", buscar_por_nombre)

    
    def buscar_solicitudes():
        nombre = combo_nombre.get()
        rut_combo = dict_nombre_rut.get(nombre, "")
        rut_entry = entry_rut.get().strip()
        if not rut_entry and rut_combo:
            entry_rut.delete(0, tk.END)
            entry_rut.insert(0, rut_combo)
        mostrar_vista_previa()

    btn_buscar.configure(command=buscar_solicitudes)

    def sumar_dias_habiles(fecha_inicio, dias_habiles):
        dias_agregados = 0
        fecha = fecha_inicio
        while dias_agregados < dias_habiles:
            if fecha.weekday() < 5:  # 0 = lunes, 4 = viernes
                dias_agregados += 1
                if dias_agregados == dias_habiles:
                    break
            fecha += timedelta(days=1)
        return fecha
    
    def sumar_dias_corridos_inclusivo(fecha_inicio, dias_corridos):
        # Ej.: lunes + 5 corridos (incluyendo lunes) => viernes
        return fecha_inicio + timedelta(days=dias_corridos - 1)

        # ==== Multiselección para eliminar (helpers) ====
    selected_ids = set()
    checkbox_vars = {}

    def actualizar_boton_eliminar():
        count = len(selected_ids)
        try:
            btn_eliminar_multi.configure(
                text=f"🗑️ Eliminar seleccionados ({count})",
                state=("normal" if count > 0 else "disabled")
            )
        except NameError:
            pass

    def on_toggle_checkbox(id_, checked):
        if checked:
            selected_ids.add(id_)
        else:
            selected_ids.discard(id_)
        actualizar_boton_eliminar()

    def seleccionar_todo():
        marcar = not all(var.get() for var in checkbox_vars.values())
        for id_, var in checkbox_vars.items():
            var.set(marcar)
            if marcar:
                selected_ids.add(id_)
            else:
                selected_ids.discard(id_)
        actualizar_boton_eliminar()

    def eliminar_seleccionados():
        if not selected_ids:
            return
        if not messagebox.askyesno("Confirmar", f"¿Eliminar {len(selected_ids)} registro(s) seleccionado(s)?"):
            return
        conn = sqlite3.connect("reloj_control.db")
        cur = conn.cursor()
        placeholders = ",".join("?" for _ in selected_ids)
        cur.execute(f"DELETE FROM dias_libres WHERE id IN ({placeholders})", tuple(selected_ids))
        conn.commit()
        conn.close()
        cant = len(selected_ids)
        selected_ids.clear()
        checkbox_vars.clear()
        mostrar_vista_previa()
        actualizar_boton_eliminar()
        messagebox.showinfo("Eliminado", f"Se eliminaron {cant} registro(s).") 

    # --- TABLA PREVIEW ---
    tabla_preview = ctk.CTkScrollableFrame(frame)
    tabla_preview.pack(fill="both", expand=True, pady=10)

    # Barra de acciones de multiselección (crear una sola vez aquí)
    acciones_sel_frame = ctk.CTkFrame(frame, fg_color="transparent")
    acciones_sel_frame.pack(fill="x", pady=(4, 6))

    btn_select_all = ctk.CTkButton(acciones_sel_frame, text="Seleccionar todo",
                                command=seleccionar_todo, width=140)
    btn_select_all.pack(side="left", padx=(0, 8))

    btn_eliminar_multi = ctk.CTkButton(
        acciones_sel_frame, text="🗑️ Eliminar seleccionados (0)",
        fg_color="red", state="disabled", command=eliminar_seleccionados, width=220
    )
    btn_eliminar_multi.pack(side="left") 

    actualizar_boton_eliminar()  

    
    
    # ========== FUNCIÓN PARA AUTO-CÁLCULO DE FECHAS ==========
    def actualizar_fecha_hasta_auto(event=None):
        tipo = combo_tipo_permiso.get()
        dias = dias_por_permiso.get(tipo)
        fecha_inicio_str = entry_fecha_desde.get().strip()

        # Si el usuario desactiva el check, no tocamos la fecha hasta
        if not auto_calc_var.get():
            return

        if dias and fecha_inicio_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, "%d/%m/%Y")

                if "días hábiles" in tipo.lower():
                    fecha_fin = sumar_dias_habiles(fecha_inicio, dias)
                else:
                    fecha_fin = sumar_dias_corridos_inclusivo(fecha_inicio, dias)

                entry_fecha_hasta.configure(state="normal")
                entry_fecha_hasta.delete(0, tk.END)
                entry_fecha_hasta.insert(0, fecha_fin.strftime("%d/%m/%Y"))
                entry_fecha_hasta.configure(state="disabled")
            except ValueError:
                entry_fecha_hasta.configure(state="normal")
                entry_fecha_hasta.delete(0, tk.END)
        else:
            entry_fecha_hasta.configure(state="normal")
            entry_fecha_hasta.delete(0, tk.END)



    entry_fecha_desde.bind("<FocusOut>", actualizar_fecha_hasta_auto)
    entry_fecha_desde.bind("<Return>", actualizar_fecha_hasta_auto)

    def guardar_dias_admin():
        rut = entry_rut.get().strip()
        fecha_inicio_str = entry_fecha_desde.get().strip()
        fecha_fin_str = entry_fecha_hasta.get().strip()
        motivo_seleccionado = combo_tipo_permiso.get()

        if motivo_seleccionado == "Elija tipo de Solicitud o Permiso":
            messagebox.showerror("Error", "Debes elegir un tipo de solicitud o permiso.")
            return

        if motivo_seleccionado == "Otro (especificar)":
            motivo = entry_motivo_otro.get().strip() or motivo_seleccionado
        elif motivo_seleccionado == "Cometido de Servicio":
            detalle = entry_motivo_otro.get().strip()
            motivo = f"Cometido de Servicio - {detalle}" if detalle else motivo_seleccionado
        else:
            motivo = motivo_seleccionado

        # Autocompletar 'hasta' si falta
        if fecha_inicio_str and not fecha_fin_str:
            if motivo_seleccionado == "Día Administrativo":
                fecha_fin_str = fecha_inicio_str
            else:
                dias = dias_por_permiso.get(motivo_seleccionado)
                if dias:
                    try:
                        fecha_inicio_tmp = datetime.strptime(fecha_inicio_str, "%d/%m/%Y")
                        if "días hábiles" in motivo_seleccionado.lower():
                            fecha_fin_tmp = sumar_dias_habiles(fecha_inicio_tmp, dias)
                        else:
                            fecha_fin_tmp = sumar_dias_corridos_inclusivo(fecha_inicio_tmp, dias)
                        fecha_fin_str = fecha_fin_tmp.strftime("%d/%m/%Y")
                    except ValueError:
                        pass  # se valida abajo

        # Validaciones
        if not rut or not fecha_inicio_str or not fecha_fin_str:
            messagebox.showerror("Error", "Debes ingresar todos los campos.")
            return

        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, "%d/%m/%Y")
            fecha_fin = datetime.strptime(fecha_fin_str, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Error", "Fechas inválidas. Usa formato dd/mm/aaaa.")
            return

        if fecha_fin < fecha_inicio:
            messagebox.showerror("Error", "La fecha final no puede ser anterior a la inicial.")
            return


        conn = sqlite3.connect("reloj_control.db")
        cur = conn.cursor()
        dias_registrados = 0
        fecha_actual = fecha_inicio
        while fecha_actual <= fecha_fin:
            # Saltar fines de semana si el permiso es de días hábiles
            if "días hábiles" in motivo_seleccionado.lower() and fecha_actual.weekday() >= 5:
                fecha_actual += timedelta(days=1)
                continue

            fecha_str = fecha_actual.strftime("%Y-%m-%d")
            cur.execute("SELECT COUNT(*) FROM dias_libres WHERE rut = ? AND fecha = ?", (rut, fecha_str))
            if cur.fetchone()[0] == 0:
                cur.execute("INSERT INTO dias_libres (rut, fecha, motivo) VALUES (?, ?, ?)", (rut, fecha_str, motivo))
                dias_registrados += 1
            fecha_actual += timedelta(days=1)

        conn.commit()
        conn.close()
        messagebox.showinfo("Éxito", f"{dias_registrados} día(s) guardado(s) correctamente.")
        mostrar_vista_previa()

    def limpiar_campos():
        entry_rut.delete(0, tk.END)
        entry_fecha_desde.delete(0, tk.END)
        entry_fecha_hasta.delete(0, tk.END)
        combo_tipo_permiso.set("Elija tipo de Solicitud o Permiso")
        entry_motivo_otro.delete(0, tk.END)
        entry_motivo_otro.pack_forget()
        combo_nombre.set("Buscar por Nombre")
        limpiar_placeholder(None)
        label_resumen_admin.configure(text="")
        label_resumen_extras.configure(text="", text_color="red")

        for widget in tabla_preview.winfo_children():
            widget.destroy()

        if auto_calc_var.get():
            entry_fecha_hasta.configure(state="disabled")
        else:
            entry_fecha_hasta.configure(state="normal")
        selected_ids.clear()
        checkbox_vars.clear()
        actualizar_boton_eliminar()

    # --- BOTONES GUARDAR Y LIMPIAR ---
    botones_frame = ctk.CTkFrame(frame, fg_color="transparent")
    botones_frame.pack(pady=(10, 10))

    ctk.CTkButton(
        botones_frame,
        text="Guardar Nueva Solicitud",
        command=guardar_dias_admin,
        width=180
    ).pack(side="left", padx=10)

    ctk.CTkButton(
        botones_frame,
        text="⬇️ Exportar informe",
        command=exportar_informe,
        width=170
    ).pack(side="left", padx=10)

    ctk.CTkButton(
        botones_frame,
        text="🧹 Limpiar Formulario",
        fg_color="gray",
        command=limpiar_campos,
        width=160
    ).pack(side="left", padx=10)



    # ========== FUNCIÓN PARA CAMBIOS DE PERMISO ==========
    def actualizar_visibilidad_entry_otro(choice):
        entry_fecha_desde.configure(state="normal")
        entry_fecha_hasta.configure(state="normal")

        if choice == "Cometido de Servicio":
            label_info_cometido.configure(
                text="⚠ Este permiso simula una jornada completa. No se requiere hora, solo el motivo."
            )
            entry_motivo_otro.pack(pady=(0, 10))
            entry_fecha_desde.delete(0, tk.END)
            entry_fecha_hasta.delete(0, tk.END)
        elif choice == "Otro (especificar)":
            label_info_cometido.configure(text="")
            entry_motivo_otro.pack(pady=(0, 10))
            entry_fecha_desde.delete(0, tk.END)
            entry_fecha_hasta.delete(0, tk.END)
        elif choice == "Día Administrativo":
            label_info_cometido.configure(text="")
            entry_motivo_otro.pack_forget()
            entry_fecha_desde.delete(0, tk.END)
            entry_fecha_hasta.delete(0, tk.END)
        elif choice in dias_por_permiso:
            label_info_cometido.configure(text="")
            entry_motivo_otro.pack_forget()
            # No borres fechas aquí
        else:
            label_info_cometido.configure(text="")
            entry_motivo_otro.pack_forget()
            entry_fecha_desde.delete(0, tk.END)
            entry_fecha_hasta.delete(0, tk.END)

        if auto_calc_var.get():
            entry_fecha_hasta.configure(state="disabled")
        else:
            entry_fecha_hasta.configure(state="normal")

        actualizar_fecha_hasta_auto()
      
    combo_tipo_permiso.configure(command=actualizar_visibilidad_entry_otro)  

    def _on_toggle_auto():
        if auto_calc_var.get():
            entry_fecha_hasta.configure(state="disabled")
            actualizar_fecha_hasta_auto()
        else:
            entry_fecha_hasta.configure(state="normal")

    chk_auto.configure(command=_on_toggle_auto)
  

    # ========== FUNCIONES PRINCIPALES ==========
    def mostrar_vista_previa():
        for widget in tabla_preview.winfo_children():
            widget.destroy()
        # limpiar selección previa
        selected_ids.clear()
        checkbox_vars.clear()
        actualizar_boton_eliminar()

        rut = entry_rut.get().strip()
        if not rut:
            label_resumen_admin.configure(text="", text_color="green")
            label_resumen_extras.configure(text="", text_color="red")
            return
        anio_actual = datetime.now().year
        conn = sqlite3.connect("reloj_control.db")
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM dias_libres 
            WHERE rut = ? 
            AND strftime('%Y', fecha) = ? 
            AND motivo = 'Día Administrativo'
        """, (rut, str(anio_actual)))
        total_admin = cur.fetchone()[0]
        MAX_ADMIN_ANUAL = 6
        color = "green" if total_admin < MAX_ADMIN_ANUAL else ("orange" if total_admin == MAX_ADMIN_ANUAL else "red")
        texto = f"📅 Días administrativos solicitados este año: {total_admin} / {MAX_ADMIN_ANUAL} disponibles"
        label_resumen_admin.configure(text=texto, text_color=color)

        cur.execute("""
            SELECT COUNT(*) FROM dias_libres 
            WHERE rut = ? 
            AND strftime('%Y', fecha) = ? 
            AND motivo != 'Día Administrativo'
        """, (rut, str(anio_actual)))
        total_extras = cur.fetchone()[0]
        texto_extras = f"🧾 Permisos extras este año: {total_extras}"
        label_resumen_extras.configure(text=texto_extras)

        cur.execute("SELECT id, fecha, motivo FROM dias_libres WHERE rut = ? ORDER BY fecha", (rut,))
        registros = cur.fetchall()
        conn.close()

        headers = ["Sel.", "Fecha", "Motivo", "Guardar", "Eliminar"]
        for i, header in enumerate(headers):
            ctk.CTkLabel(tabla_preview, text=header, font=("Arial", 13, "bold")).grid(row=0, column=i, padx=10, pady=5)

        for idx, (id_, fecha, motivo) in enumerate(registros, start=1):
            # checkbox de selección
            var = tk.BooleanVar(value=False)
            chk = ctk.CTkCheckBox(tabla_preview, text="", variable=var,
                                command=lambda i=id_, v=var: on_toggle_checkbox(i, v.get()))
            chk.grid(row=idx, column=0, padx=6, pady=2)
            checkbox_vars[id_] = var

            fecha_legible = datetime.strptime(fecha, "%Y-%m-%d").strftime("%d/%m/%Y")
            ctk.CTkLabel(tabla_preview, text=fecha_legible).grid(row=idx, column=1, padx=10, pady=2)

            entry_motivo_edit = ctk.CTkEntry(tabla_preview, width=400, font=("Arial", 14))
            entry_motivo_edit.insert(0, motivo or "")
            entry_motivo_edit.grid(row=idx, column=2, padx=10, pady=2)

            ctk.CTkButton(tabla_preview, text="💾", width=30, fg_color="green",
                        command=lambda i=id_, e=entry_motivo_edit: actualizar_motivo(i, e)).grid(row=idx, column=3, padx=5)

            ctk.CTkButton(tabla_preview, text="❌", width=30, fg_color="red",
                        command=lambda i=id_: eliminar_dia_admin(i)).grid(row=idx, column=4, padx=5)


    

    def actualizar_motivo(id_, entry_widget):
        nuevo_motivo = entry_widget.get().strip()
        conn = sqlite3.connect("reloj_control.db")
        cur = conn.cursor()
        cur.execute("UPDATE dias_libres SET motivo = ? WHERE id = ?", (nuevo_motivo, id_))
        conn.commit()
        conn.close()
        messagebox.showinfo("Actualizado", "Motivo actualizado correctamente.")

    def eliminar_dia_admin(id_):
        confirm = messagebox.askyesno("Confirmar", "¿Deseas eliminar este día administrativo?")
        if not confirm:
            return
        conn = sqlite3.connect("reloj_control.db")
        cur = conn.cursor()
        cur.execute("DELETE FROM dias_libres WHERE id = ?", (id_,))
        conn.commit()
        conn.close()
        mostrar_vista_previa()
        messagebox.showinfo("Eliminado", "Día administrativo eliminado correctamente.")

    