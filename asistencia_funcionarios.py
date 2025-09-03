# asistencia_funcionarios.py
import os
import sqlite3
import calendar
import datetime as dt
import tkinter as tk
from tkinter import ttk, messagebox
import customtkinter as ctk

# Calendario opcional
try:
    from tkcalendar import DateEntry
    HAS_TKCAL = True
except Exception:
    HAS_TKCAL = False

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None

# Email (opcional)
import smtplib
from email.message import EmailMessage
import mimetypes
import traceback
import sys
import subprocess


MESES_ES = {
    1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
    7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"
}

def _downloads_dir():
    home = os.path.expanduser("~")
    cand = os.path.join(home, "Downloads")
    if os.path.isdir(cand): return cand
    xdg = os.environ.get("XDG_DOWNLOAD_DIR")
    if xdg and os.path.isdir(xdg): return xdg
    return home

def _abrir_archivo(path):
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception as e:
        messagebox.showinfo("Archivo generado", f"Se guardó en:\n{path}\n\nNo se pudo abrir automáticamente: {e}")

def _mes_range(year: int, month: int):
    # días exactos del mes (calendar maneja años bisiestos)
    last_day = calendar.monthrange(year, month)[1]
    start = dt.date(year, month, 1)
    end = dt.date(year, month, last_day)
    return start, end, last_day

def _leer_funcionarios(con):
    cur = con.cursor()
    cur.execute("""
        SELECT rut, (nombre || ' ' || apellido) AS nombre_completo
        FROM trabajadores
        ORDER BY nombre_completo COLLATE NOCASE;
    """)
    return cur.fetchall()  # [(rut, nombre_completo), ...]

def _cargar_registros(con, rut: str | None, f_ini: dt.date, f_fin: dt.date):
    """
    Devuelve filas normalizadas: (rut, nombre, fecha, tipo, hora)
    combinando eventos (tipo/hora) y columnas (hora_ingreso/hora_salida).
    """
    cur = con.cursor()
    params = [f_ini.strftime("%Y-%m-%d"), f_fin.strftime("%Y-%m-%d")]
    filtro = ""
    if rut:
        filtro = " AND r.rut = ? "
        params.append(rut)
    cur.execute(f"""
        SELECT 
            r.rut,
            COALESCE(t.nombre || ' ' || t.apellido, r.nombre, r.rut) AS nombre,
            r.fecha,
            MAX(CASE WHEN LOWER(IFNULL(r.tipo,'')) LIKE 'ing%%' THEN r.hora END) AS evt_ingreso,
            MAX(CASE WHEN LOWER(IFNULL(r.tipo,'')) LIKE 'sal%%' THEN r.hora END) AS evt_salida,
            MAX(NULLIF(TRIM(IFNULL(r.hora_ingreso,'')), '')) AS col_ingreso,
            MAX(NULLIF(TRIM(IFNULL(r.hora_salida,'')), '')) AS col_salida
        FROM registros r
        LEFT JOIN trabajadores t ON t.rut = r.rut
        WHERE r.fecha BETWEEN ? AND ? {filtro}
        GROUP BY r.rut, r.fecha
        ORDER BY r.rut, r.fecha;
    """, params)
    agg = cur.fetchall()

    norm_rows = []
    for rutx, nombre, fecha, evt_ing, evt_sal, col_ing, col_sal in agg:
        hora_ing = col_ing or evt_ing
        hora_sal = col_sal or evt_sal
        if hora_ing:
            norm_rows.append((rutx, nombre, fecha, "ingreso", hora_ing))
        if hora_sal:
            norm_rows.append((rutx, nombre, fecha, "salida", hora_sal))
    return norm_rows

def _resumen_por_rut(rows):
    from collections import defaultdict
    data = defaultdict(lambda: {
        "nombre": "",
        "fechas": set(),
        "por_fecha": {},  # fecha -> {"ingreso": hh:mm, "salida": hh:mm}
    })
    for rut, nombre, fecha, tipo, hora in rows:
        d = data[rut]
        d["nombre"] = nombre
        d["fechas"].add(fecha)
        if fecha not in d["por_fecha"]:
            d["por_fecha"][fecha] = {"ingreso": None, "salida": None}
        if tipo.lower().startswith("ing"):
            d["por_fecha"][fecha]["ingreso"] = (hora[:5] if hora else None)
        elif tipo.lower().startswith("sal"):
            d["por_fecha"][fecha]["salida"] = (hora[:5] if hora else None)
    return data

def _style_dark_treeview():
    # Tema oscuro para ttk.Treeview
    style = ttk.Style()
    try:
        style.theme_use('clam')
    except Exception:
        pass
    style.configure(
        "Dark.Treeview",
        background="#1e1e1e",
        foreground="#e5e5e5",
        fieldbackground="#1e1e1e",
        rowheight=26,
        bordercolor="#333333",
        borderwidth=0
    )
    style.configure("Dark.Treeview.Heading",
        background="#2a2a2a",
        foreground="#d0d0d0",
        relief="flat"
    )
    style.map("Dark.Treeview",
        background=[("selected","#094771")],
        foreground=[("selected","#ffffff")]
    )
    return "Dark.Treeview", "Dark.Treeview.Heading"

def _construir_tabla_preview(parent):
    tv_style, tv_head = _style_dark_treeview()
    cols = ("nombre", "rut", "rango", "total_dias_mes", "dias_trabajados", "observaciones")
    tree = ttk.Treeview(parent, columns=cols, show="headings", height=12, style=tv_style)
    headers = {
        "nombre": "Nombre",
        "rut": "RUT",
        "rango": "Rango",
        "total_dias_mes": "Total días mes",
        "dias_trabajados": "Días trabajados",
        "observaciones": "Observaciones (completo/incompleto)"
    }
    for c in cols:
        tree.heading(c, text=headers[c])
        tree.column(c, width=150 if c not in ("observaciones","rango") else (520 if c=="observaciones" else 180), anchor="w")
    tree.pack(fill="both", expand=True)
    return tree

def abrir_asistencia(app_root, db_path: str, *, default_todos: bool = False):
    TopLevelCls = getattr(ctk, "CTkToplevel", None) or tk.Toplevel
    win = TopLevelCls(app_root)
    win.title("Asistencia de Funcionarios")
    try:
        win.resizable(True, True); win.transient(app_root); win.grab_set()
    except Exception: pass

    cont = ctk.CTkFrame(win)
    cont.pack(fill="both", expand=True, padx=12, pady=12)

    # ---------- fila 1: buscador ----------
    fila1 = ctk.CTkFrame(cont)
    fila1.pack(fill="x", padx=6, pady=(6, 10))

    ctk.CTkLabel(fila1, text="Buscar").grid(row=0, column=0, padx=(0, 8), pady=6, sticky="w")
    entry_buscar = ctk.CTkEntry(fila1, placeholder_text="Nombre o RUT")
    entry_buscar.grid(row=0, column=1, padx=(0, 10), pady=6, sticky="ew")

    ctk.CTkLabel(fila1, text="Funcionario").grid(row=0, column=2, padx=(0, 8), pady=6, sticky="w")
    combo_func = ctk.CTkComboBox(fila1, values=["Cargando..."])
    combo_func.grid(row=0, column=3, padx=(0, 8), pady=6, sticky="ew")

    var_todos = tk.BooleanVar(value=bool(default_todos))
    chk_todos = ctk.CTkCheckBox(fila1, text="Todos", variable=var_todos)
    chk_todos.grid(row=0, column=4, padx=(4, 0), pady=6, sticky="w")

    fila1.grid_columnconfigure(1, weight=1)
    fila1.grid_columnconfigure(3, weight=1)

    con = sqlite3.connect(db_path)
    funcionarios = _leer_funcionarios(con)
    listado = [f"{n} | {r}" for (r, n) in funcionarios]
    combo_func.configure(values=listado if listado else ["(Sin registros)"])
    if listado: combo_func.set(listado[0])

    def filtrar():
        q = entry_buscar.get().strip().lower()
        vals = [x for x in listado if q in x.lower()] if q else listado
        combo_func.configure(values=vals if vals else ["(Sin coincidencias)"])
        if vals: combo_func.set(vals[0])
    entry_buscar.bind("<KeyRelease>", lambda e: filtrar())

    # ---------- fila 2: rango / mes ----------
    fila2 = ctk.CTkFrame(cont)
    fila2.pack(fill="x", padx=6, pady=(0, 10))

    modo_var = tk.StringVar(value="rango")
    r1 = ctk.CTkRadioButton(fila2, text="Rango de fechas", variable=modo_var, value="rango")
    r2 = ctk.CTkRadioButton(fila2, text="Mes completo", variable=modo_var, value="mensual")
    r1.grid(row=0, column=0, padx=(0, 12), pady=6, sticky="w")
    r2.grid(row=0, column=1, padx=(0, 12), pady=6, sticky="w")

    ctk.CTkLabel(fila2, text="Desde").grid(row=1, column=0, padx=(0, 8), pady=6, sticky="e")
    if HAS_TKCAL:
        entry_desde = DateEntry(fila2, date_pattern="yyyy-mm-dd", locale="es_CL")
        entry_hasta = DateEntry(fila2, date_pattern="yyyy-mm-dd", locale="es_CL")
    else:
        entry_desde = tk.Entry(fila2, width=12)
        entry_hasta = tk.Entry(fila2, width=12)
        entry_desde.insert(0, dt.date.today().strftime("%Y-%m-%d"))
        entry_hasta.insert(0, dt.date.today().strftime("%Y-%m-%d"))
    entry_desde.grid(row=1, column=1, padx=(0, 16), pady=6, sticky="w")

    ctk.CTkLabel(fila2, text="Hasta").grid(row=1, column=2, padx=(0, 8), pady=6, sticky="e")
    entry_hasta.grid(row=1, column=3, padx=(0, 16), pady=6, sticky="w")

    ctk.CTkLabel(fila2, text="Año").grid(row=1, column=4, padx=(12, 8), pady=6, sticky="e")
    combo_anio = ctk.CTkComboBox(fila2, values=[str(y) for y in range(dt.date.today().year - 3, dt.date.today().year + 2)])
    combo_anio.set(str(dt.date.today().year))
    combo_anio.grid(row=1, column=5, padx=(0, 16), pady=6, sticky="w")

    ctk.CTkLabel(fila2, text="Mes").grid(row=1, column=6, padx=(0, 8), pady=6, sticky="e")
    meses = [f"{i:02d}" for i in range(1, 13)]
    combo_mes = ctk.CTkComboBox(fila2, values=meses)
    combo_mes.set(f"{dt.date.today().month:02d}")
    combo_mes.grid(row=1, column=7, padx=(0, 0), pady=6, sticky="w")

    btn_prev = ctk.CTkButton(fila2, text="Previsualizar")
    btn_prev.grid(row=1, column=8, padx=(16, 0), pady=6, sticky="w")

    # ---------- tabla ----------
    tabla_frame = ctk.CTkFrame(cont)
    tabla_frame.pack(fill="both", expand=True, padx=6, pady=(0, 10))
    tabla = _construir_tabla_preview(tabla_frame)

    # ---------- botones inferiores ----------
    fila3 = ctk.CTkFrame(cont); fila3.pack(fill="x", padx=6, pady=(0, 6))
    btn_descargar = ctk.CTkButton(fila3, text="Descargar")
    btn_enviar = ctk.CTkButton(fila3, text="Enviar por correo")
    btn_cancelar = ctk.CTkButton(fila3, text="Cancelar", fg_color="gray", command=win.destroy)
    btn_descargar.pack(side="left", padx=4); btn_enviar.pack(side="left", padx=4); btn_cancelar.pack(side="right", padx=4)

    # ---------- lógica ----------
    def _obtener_rut_sel():
        if var_todos.get(): return None
        val = combo_func.get().strip()
        if " | " in val:
            _, rut = val.split(" | ", 1)
            return rut.strip()
        for r, n in funcionarios:
            if val == f"{n} | {r}":
                return r
        return None

    def _rango_fechas():
        if modo_var.get() == "rango":
            if HAS_TKCAL:
                f_ini = dt.datetime.strptime(entry_desde.get(), "%Y-%m-%d").date()
                f_fin = dt.datetime.strptime(entry_hasta.get(), "%Y-%m-%d").date()
            else:
                f_ini = dt.datetime.strptime(entry_desde.get().strip(), "%Y-%m-%d").date()
                f_fin = dt.datetime.strptime(entry_hasta.get().strip(), "%Y-%m-%d").date()
            return f_ini, f_fin, f"{f_ini} a {f_fin}", None, (f_fin - f_ini).days + 1
        else:
            y = int(combo_anio.get()); m = int(combo_mes.get())
            f_ini, f_fin, last_day = _mes_range(y, m)
            etiqueta = f"{y} - {MESES_ES[m]}"
            return f_ini, f_fin, etiqueta, (y, m), last_day

    def previsualizar():
        try:
            rut_sel = _obtener_rut_sel()
            f_ini, f_fin, etiqueta_rango, ym, total_dias = _rango_fechas()
            rows = _cargar_registros(con, rut_sel, f_ini, f_fin)
            tabla.delete(*tabla.get_children())

            agg = _resumen_por_rut(rows)

            if rut_sel is None and not agg:
                # todos sin movimientos
                for r, n in funcionarios:
                    tabla.insert("", "end", values=(n, r, etiqueta_rango, total_dias, 0, "Sin movimientos"))
                return

            for rutx, info in (agg.items() if agg else []):
                nombre = info["nombre"]
                dias_trab = len(info["fechas"])
                completos = []
                incompletos = []
                for f, p in sorted(info["por_fecha"].items()):
                    if p["ingreso"] and p["salida"]:
                        completos.append(f)
                    else:
                        falt = []
                        if not p["ingreso"]: falt.append("ingreso")
                        if not p["salida"]:  falt.append("salida")
                        incompletos.append(f"{f} (sin {', '.join(falt)})")
                obs = []
                if completos:   obs.append(f"Días completos: {len(completos)}")
                if incompletos: obs.append("Incompletos: " + "; ".join(incompletos[:6]) + (" ..." if len(incompletos) > 6 else ""))
                if not obs:
                    obs = ["Sin movimientos"] if dias_trab == 0 else ["Registros sin detalle"]
                tabla.insert("", "end", values=(nombre, rutx, etiqueta_rango, total_dias, dias_trab, " | ".join(obs)))

            if rut_sel and not tabla.get_children():
                f = next(((r, n) for (r, n) in funcionarios if r == rut_sel), None)
                if f:
                    _, n = f
                    tabla.insert("", "end", values=(n, rut_sel, etiqueta_rango, total_dias, 0, "Sin movimientos"))

        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Previsualizar", f"Error al calcular asistencia:\n{e}")

    btn_prev.configure(command=previsualizar)

    def descargar():
        if Workbook is None:
            messagebox.showerror("Descargar", "openpyxl no está instalado.")
            return
        filas = tabla.get_children()
        if not filas:
            messagebox.showinfo("Descargar", "No hay datos para exportar.")
            return
        wb = Workbook(); ws = wb.active; ws.title = "Asistencia"
        headers = ["Nombre", "RUT", "Rango", "Total días mes", "Días trabajados", "Observaciones"]
        ws.append(headers)
        for item in filas:
            ws.append(list(tabla.item(item, "values")))
        for ci in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(ci)].width = 26 if ci != len(headers) else 70
        try:
            desc = _downloads_dir()
            nombre = f"Asistencia_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            path = os.path.join(desc, nombre)
            wb.save(path); _abrir_archivo(path)
        except Exception as e:
            traceback.print_exc(); messagebox.showerror("Descargar", f"No se pudo guardar/abrir:\n{e}")
    btn_descargar.configure(command=descargar)

    def enviar_correo():
        filas = tabla.get_children()
        if not filas: return messagebox.showinfo("Correo","No hay datos para enviar.")
        if Workbook is None: return messagebox.showerror("Correo","openpyxl no está instalado.")
        try:
            wb = Workbook(); ws = wb.active; ws.title = "Asistencia"
            headers = ["Nombre", "RUT", "Rango", "Total días mes", "Días trabajados", "Observaciones"]
            ws.append(headers)
            for it in filas:
                ws.append(list(tabla.item(it, "values")))
            tmp_dir = _downloads_dir()
            path = os.path.join(tmp_dir, f"Asistencia_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb.save(path)
        except Exception as e:
            traceback.print_exc(); return messagebox.showerror("Correo", f"No se pudo preparar el archivo:\n{e}")

        top = tk.Toplevel(win); top.title("Enviar por correo")
        tk.Label(top, text="Para:").grid(row=0, column=0, padx=8, pady=8, sticky="e")
        entry_to = tk.Entry(top, width=40); entry_to.grid(row=0, column=1, padx=8, pady=8, sticky="w")

        def _do_send():
            to = entry_to.get().strip()
            if not to: return messagebox.showerror("Correo","Debes indicar un destinatario.")
            try:
                SMTP_HOST = os.environ.get("SMTP_HOST", ""); SMTP_PORT = int(os.environ.get("SMTP_PORT","587"))
                SMTP_USER = os.environ.get("SMTP_USER","");  SMTP_PASS = os.environ.get("SMTP_PASS","")
                if not SMTP_HOST or not SMTP_USER or not SMTP_PASS:
                    return messagebox.showwarning("Correo","Falta configuración SMTP (env vars).")
                msg = EmailMessage()
                msg["Subject"]="Asistencia de Funcionarios"; msg["From"]=SMTP_USER; msg["To"]=to
                msg.set_content("Adjunto asistencia generada desde BioAccess.")
                ctype,_ = mimetypes.guess_type(path); maintype,subtype=(ctype.split("/",1) if ctype else ("application","octet-stream"))
                with open(path,"rb") as f: msg.add_attachment(f.read(), maintype=maintype, subtype=subtype, filename=os.path.basename(path))
                with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as s:
                    s.starttls(); s.login(SMTP_USER, SMTP_PASS); s.send_message(msg)
                messagebox.showinfo("Correo","Enviado correctamente."); top.destroy()
            except Exception as e:
                traceback.print_exc(); messagebox.showerror("Correo", f"No se pudo enviar:\n{e}")

        tk.Button(top, text="Cancelar", command=top.destroy).grid(row=1, column=0, padx=8, pady=(0,8), sticky="ew")
        tk.Button(top, text="Enviar", command=_do_send).grid(row=1, column=1, padx=8, pady=(0,8), sticky="ew")
        top.grab_set(); top.transient(win); entry_to.focus_set()

    btn_enviar.configure(command=enviar_correo)

    # Centrar
    app_root.update_idletasks()
    w,h = 980, 620
    x = app_root.winfo_x() + (app_root.winfo_width() // 2) - (w // 2)
    y = app_root.winfo_y() + (app_root.winfo_height() // 2) - (h // 2)
    win.geometry(f"{w}x{h}+{max(x,0)}+{max(y,0)}")
